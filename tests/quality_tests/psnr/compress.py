from PIL import Image
import sys
from io import BytesIO
import cv2
import numpy as np
import pandas as pd
from openpyxl import load_workbook


NAME = sys.argv[1]
Q = sys.argv[2]
IMAGE1 = sys.argv[3]
#IMAGE2 = sys.argv[4]



psnr_dic = {
	"IMAGE":None,
	"Q=95":None,
	"Q=85":None,
	"Q=75":None,
	"Q=65":None,
}


def initq():
	qualities = Q.split("-")
	return qualities

def compressImage(path,q):
	im = Image.open(path)
	im.save("compressed_"+NAME+str(q)+".jpg",
			optimize = True, 
			quality = int(q))
	return im

def psnr(im1,im2,q):
	img1 = Image.open(im1)
	#img2 = Image.open(im2)
	p = cv2.PSNR(np.array(img1),np.array(im2))
	print(p)
	p = (round(p, 2))
	return p

def psnrPerQ():
	quality_list = initq()
	for quality in quality_list:
		cimg = compressImage(IMAGE1,quality)
		#p = psnr(IMAGE1,cimg,quality)
		#psnr_dic["IMAGE"] = IMAGE1
		#q_measure = "Q=" + quality
		#psnr_dic[q_measure] = [str(p)]

def writeToExcel():
	df = pd.DataFrame(psnr_dic, columns = ['IMAGE', 'Q=95',"Q=85","Q=75","Q=65"])
	book = load_workbook("img_psnr.xlsx")
	writer = pd.ExcelWriter("img_psnr.xlsx", engine='openpyxl')
	writer.book = book
	writer.sheets = {ws.title: ws for ws in book.worksheets}
	df.to_excel(writer, startrow=writer.sheets['Sheet1'].max_row, index = False,header= False)
	writer.save()

def getDic():
	print(psnr_dic)

if __name__ == '__main__':
	psnrPerQ()
	#writeToExcel()
	#getDic()
	